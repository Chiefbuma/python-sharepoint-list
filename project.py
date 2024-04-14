from sharepoint import SharePoint
from openpyxl import Workbook
import streamlit as st
import pandas as pd

# get clients sharepoint list
clients = SharePoint().connect_to_list(ls_name='Maintenance Report')

# create DataFrame from clients list
df = pd.DataFrame(clients)

st.write(df)

# create excel workbook
wb = Workbook()

dest_filepath = 'maintenance_list.xlsx'

# create worksheet
ws = wb.active
ws.title = 'Maintenance'

# setting sharepoint list values to excel cells
for idx, client in enumerate(clients, 1):
    ws.cell(column=1, row=idx, value=client['Title'])
    ws.cell(column=2, row=idx, value=client['Details'])
    ws.cell(column=3, row=idx, value=client['ID'])

# save workbook
wb.save(filename=dest_filepath)
